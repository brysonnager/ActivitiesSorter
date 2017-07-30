#READ THE FOLLOWING LINES BEFORE USING!!!!

#After running, input 'checkStudentInfo()' to learn about a student's preferences
#       and current activity
#input 'check3OrBelow()' to learn about student's who didn't get either first or second choice




'''To run this program, you must install something called openpyxl,
otherwise, the program won't work. Here are instructions to do so on Mac...
1. Go to https://pypi.python.org/pypi/openpyxl
2. Note the version that you are downloading, it should be written
directly above the green "Download" button
3. Click Download
4. Open Terminal
5. Type the following statement, depending on your version. My version was
openpyxl 2.5.0a2 so I typed 'pip3 install openpyxl==2.5.0a2'
6. Press enter

Now, openpyxl should be installed.


Also, please save this file in a folder and also in that folder include
an Excel doc entitled 'activities_form.xlsx' containing the information
from the Google sheet for the week's activities.



The info will be saved in a new Excel doc called
sorted.xlsx


Press 'fn' + 'F5' to run this program!!'''






import openpyxl
import random

wb = openpyxl.load_workbook('activities_form.xlsx')
names = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(names[0])
#sheet.cell(row=4,column=6).value

day = input("Which day? (Monday, Tuesday, Wednesday, Thursday, Friday)")
day = day.title()

#first and last columns for that day
day_start_column = 0
day_end_column = 0

for i in range(1,60):
    if day_start_column == 0:
        if day in sheet.cell(row=1,column=i).value:
            day_start_column = i
    elif day_end_column == 0:
        if day not in sheet.cell(row=1,column=i).value:
            day_end_column = i-1







class Person():
    first_name = ''
    last_name = ''
    full_name = ''
    house = ''
    preferences = []
    role = ''
    def __init__(self, first_name, last_name, house, role, preferences):
        self.first_name = first_name
        self.last_name = last_name
        self.full_name = first_name + " " + last_name
        self.house = house
        self.preferences = preferences
        for i in range(len(preferences)):
            if type(preferences[i]) != int:
                preferences[i] = int(preferences[i])


#list of students
students = []
#list of staff
staff = []



#making instances of the class Person
current_row = 2
while True:
    if sheet.cell(row = current_row, column=2).value is None:
        break

    f_n = str(sheet.cell(row = current_row, column=2).value).title()
    l_n = str(sheet.cell(row = current_row, column=3).value).title()
    house = str(sheet.cell(row = current_row, column=4).value)
    role = str(sheet.cell(row = current_row, column=5).value)

    current_column = day_start_column
    preferences = []
    while current_column <= day_end_column:
        try:
            preferences.append(sheet.cell(row = current_row, column=current_column).value[0])
        except:
            preferences.append(sheet.cell(row = current_row, column=current_column).value)
        finally:
            current_column = current_column + 1
        
    if role == "Participant":
        students.append(Person(f_n, l_n, house, role, preferences))
    else:
        staff.append(Person(f_n, l_n, house, role, preferences))
    
    current_row = current_row + 1


''' AT THIS POINT, all the Person objects have been made and it is time to begin sorting! '''

activities = []

class Activity():
    name = ''
    members = []
    open = 0
    cancelled = 0

    def __init__(self, entry): #entry is the name
        self.name = entry
        self.members = []
        self.open = 1

    def bubbleSortList(self):
        for i in range(len(self.members)):
            for j in range(len(self.members) - 1):
                if self.members[j].house > self.members[j+1].house:
                    temp = self.members[j]
                    self.members[j] = self.members[j+1]
                    self.members[j+1] = temp

    def restrictActivity(self, num):
        self.open = 0    #this activity is closed, nobody can be added

        print("Students deleted from this activity:") 
        while len(self.members) > num:
            random_index = random.randint(0, len(self.members))
            try:
                student = self.members[random_index]
                del self.members[random_index]
                replaced = 0
                for i in student.preferences:
                    if replaced == 0:
                        if i == 2 and activities[student.preferences.index(i)].open == 1:
                            activities[student.preferences.index(i)].members.append(student)
                            print(student.full_name)
                            replaced = 1
                if replaced == 0:
                    for i in student.preferences:
                        if replaced == 0:
                            if i == 3 and activities[student.preferences.index(i)].open == 1:
                                activities[student.preferences.index(i)].members.append(student)
                                print(student.full_name)
                                replaced = 1
                if replaced == 0:
                    for i in student.preferences:
                        if replaced == 0:
                            if i == 4 and activities[student.preferences.index(i)].open == 1:
                                activities[student.preferences.index(i)].members.append(student)
                                print(student.full_name)
                                replaced = 1
                if replaced == 0:
                    for i in student.preferences:
                        if replaced == 0:
                            if i == 5 and activities[student.preferences.index(i)].open == 1:
                                activities[student.preferences.index(i)].members.append(student)
                                print(student.full_name)
                                replaced = 1
                if replaced == 0:
                    for i in student.preferences:
                        if replaced == 0:
                            if i == 6 and activities[student.preferences.index(i)].open == 1:
                                activities[student.preferences.index(i)].members.append(student)
                                print(student.full_name)
                                replaced = 1
                    
            except:
                x = 0

    def cancelActivity(self):
        self.cancelled = 1
        self.open = 0
        for student in self.members:
            while len(self.members) > 0:
                random_index = random.randint(0, len(self.members))
                try:
                    student = self.members[random_index]
                    del self.members[random_index]
                    replaced = 0
                    for i in student.preferences:
                        if replaced == 0:
                            if i == 2 and activities[student.preferences.index(i)].open == 1:
                                activities[student.preferences.index(i)].members.append(student)
                                print(student.full_name)
                                replaced = 1
                    if replaced == 0:
                        for i in student.preferences:
                            if replaced == 0:
                                if i == 3 and activities[student.preferences.index(i)].open == 1:
                                    activities[student.preferences.index(i)].members.append(student)
                                    print(student.full_name)
                                    replaced = 1
                    if replaced == 0:
                        for i in student.preferences:
                            if replaced == 0:
                                if i == 4 and activities[student.preferences.index(i)].open == 1:
                                    activities[student.preferences.index(i)].members.append(student)
                                    print(student.full_name)
                                    replaced = 1
                    if replaced == 0:
                        for i in student.preferences:
                            if replaced == 0:
                                if i == 5 and activities[student.preferences.index(i)].open == 1:
                                    activities[student.preferences.index(i)].members.append(student)
                                    print(student.full_name)
                                    replaced = 1
                    if replaced == 0:
                        for i in student.preferences:
                            if replaced == 0:
                                if i == 6 and activities[student.preferences.index(i)].open == 1:
                                    activities[student.preferences.index(i)].members.append(student)
                                    print(student.full_name)
                                    replaced = 1
                except:
                    x = 0


counter = day_start_column
while counter <= day_end_column:
    entry = str(sheet.cell(row = 1, column=counter).value)
    activities.append(Activity(entry))
    counter = counter + 1
                                           
    
def printStatistics():
    for i in range(len(activities)):
        print(activities[i].name)
        for j in range(1, len(activities)+1):
            num = 0
            for student in students:
                if student.preferences[i] == j:
                    num = num + 1
            print("Choice #" + str(j) + "    " + str(num))




#printStatistics()







#sort by first choice
for student in students:
    num = 0

    while num < len(activities):
        if student.preferences[num] == 1:
            activities[num].members.append(student)
        num = num + 1



#sorted by first choice

for activity in activities:
    print(activity.name + " currently has " + str(len(activity.members)))


while True:
    print("What would you like to do?")
    print("*" * 40)
    print("1: Set a limit on an activity")
    print("2: Cancel an activity")
    print("3: Make the spreadsheet")
    print("4: Check for Errors! (shows students who aren't in an activity)")
    print("\t\t(It's possible that they didn't fill out the form correctly)")
    print("*" * 40)
    user_input = 0
    try:
        user_input = int(input("Please type 1, 2, 3, or 4"))
    except:
        print("You didn't type 1,2, or 3")

    if user_input == 1:
        print("*" * 40)
        print(" ")
        for i in range(0, len(activities)):
            print("Option " + str(i) + ": " + activities[i].name)
        print(" ")
        print("*" * 40)
        print(" ")
        choice = int(input("Please input your choice (0-" + str(len(activities)-1) + ")"))
        max_num = int(input("Max number for this activity?"))


        activities[choice].restrictActivity(max_num)
        #Change this activity to account for this change
        for activity in activities:
            print(activity.name + " currently has " + str(len(activity.members)))



        
    if user_input == 2:
        print("*" * 40)
        print(" ")
        for i in range(0, len(activities)):
            print("Option " + str(i) + ": " + activities[i].name)
        print(" ")
        print("*" * 40)
        print(" ")
        choice = int(input("Please input your choice (0-" + str(len(activities)-1) + ")"))
        inp = input("Type 'yes' if you would really like to cancel this activity")
        if inp.upper() == "YES":
            activities[choice].cancelActivity()
        for activity in activities:
            print(activity.name + " currently has " + str(len(activity.members)))

    

        


        
    if user_input == 4:
        not_found = []
        for student in students:
            found = 0
            for activity in activities:
                if student in activity.members:
                    found = 1
            if found == 0:
                not_found.append(student)
        if len(not_found) == 0:
            print("All students accounted for!")
        else:
            print("Students not found: ")
            for student in not_found:
                print(student.full_name)



    if user_input == 3:
        workbook = openpyxl.Workbook()
        new = workbook.active


        for activity in activities:
            activity.bubbleSortList()
        


        current_column = 2
        current_row = 2
        for activity in activities:
            new.cell(row = 1, column = current_column).value = activity.name
            current_row = 2
            num = 1
            for member in activity.members:
                #number, firstname, lastname, house
                new.cell(row=current_row, column = current_column).value = str(num)
                new.cell(row=current_row, column = current_column + 1).value = member.first_name
                new.cell(row=current_row, column = current_column + 2).value = member.last_name
                new.cell(row=current_row, column = current_column + 3).value = member.house
                current_row = current_row + 1
                num = num + 1
                
 #       new.cell(row=1, column=2).value = 2
            current_column = current_column + 5



        workbook.save('sorted.xlsx')



        print("*" * 20)
        print(" ")
        print(" ")
        print("Total number of students: " + str(len(students)))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 1:
                    choice_count = choice_count + 1
        print("Students with first choice: " + str(choice_count))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 2:
                    choice_count = choice_count + 1
        print("Students with second choice: " + str(choice_count))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 3:
                    choice_count = choice_count + 1
        print("Students with third choice: " + str(choice_count))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 4:
                    choice_count = choice_count + 1
        print("Students with fourth choice: " + str(choice_count))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 5:
                    choice_count = choice_count + 1
        print("Students with fifth choice: " + str(choice_count))
        choice_count = 0
        for x in range(0, len(activities)):
            for student in activities[x].members:
                if student.preferences[x] == 6:
                    choice_count = choice_count + 1
        print("Students with sixth choice: " + str(choice_count))
        print("If these numbers don't add up, something went wrong, please re-run the program")
        print("It is possible that the spreadsheet was filled out wrong by a student.")

        for student in students:
            num = 0
            for activity in activities:
                if student in activity.members:
                    num = 1
            if num == 0:
                print("Uh-oh, " + student.full_name + " doesn't have an activity")





        
        break
        


def checkStudentInfo():
    name = input("Input student name to check info of student")
    name = name.upper()
    name = name.strip()
    for student in students:
        if student.first_name.upper() == name or student.full_name.upper() == name or student.last_name.upper() == name:
            print(student.full_name.title() + " ")
            for i in range(len(activities)):
                print("\t" + str(student.preferences[i]))
            activity_num = 0
            found = 0
            for activity in activities:
                if student in activity.members:
                    activity_num = activities.index(activity)
                    found = 1
            if found == 1:
                print("Activity: " + activities[activity_num].name)
            else:
                print("Error: Student not found")



def check3OrBelow():
    for activity in activities:
        for student in activity.members:
            index = activities.index(activity)
            if student.preferences[index] > 2:
                print(student.full_name + " got choice number " + str(student.preferences[index]) + " (" + activity.name + ")")




















